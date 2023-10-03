from PhRegPrv import PhRegPrv
from openpyxl import load_workbook
import os
import xlrd
class Edrw:
    def __init__(self):
        pass
    def update_sources(self, regName, qtr, baseFolder, sdFile, commcode, yr):
        philippines = PhRegPrv()
        if regName == "All":
            selected_regions = philippines.get_regions()
        else:
            selected_regions = [regName]
        print(selected_regions)
        for regName in selected_regions:
            selected_provinces = philippines.get_provinces(regName)
            src = load_workbook(filename = baseFolder + "/" + sdFile)
            if regName == "Ilocos Region":
                regCode = "01"
            elif regName == "Cagayan Valley":
                regCode = "02"
            elif regName == "Central Luzon":
                regCode = "03"
            elif regName == "CALABARZON":
                regCode = "04"
            elif regName == "Bicol Region":
                regCode = "05"
            elif regName == "Western Visayas":
                regCode = "06"
            elif regName == "Central Visayas":
                regCode = "07"
            elif regName == "Eastern Visayas":
                regCode = "08"
            elif regName == "Zamboanga Peninsula":
                regCode = "09"
            elif regName == "Northern Mindanao":
                regCode = "10"
            elif regName == "Davao Region":
                regCode = "11"
            elif regName == "SOCCSKSARGEN":
                regCode = "12"
            elif regName == "NCR":
                regCode = "13"
            elif regName == "CAR":
                regCode = "14"
            elif regName == "BARMM":
                regCode = "15"
            elif regName == "Caraga":
                regCode = "16"
            elif regName == "MIMAROPA Region":
                regCode = "17"

            if commcode == '08':
                src_active = src["native B"] # Get Province Names from this worksheet
                comm_arr = ["native","gamefowl","broiler","layer"]
                srv_arr = [" B"," C","_T"]
                if selected_provinces:
                    for province in selected_provinces:
                        #print(f"({province}")
                        srcFile = str(commcode) + ' ' + str(province) + '_' + str(yr)
                        fileExt = '.xlsm'
                        fpath = str(baseFolder) + '/Sources/' + str(regName) + '/' + str(srcFile) + str(fileExt)
                        if os.path.isfile(fpath) == True:
                            for row in src_active.iter_rows(min_row=17, min_col=1, max_row=135, max_col=1):
                                for cell in row:
                                    if cell.value == province:
                                        print(f"{province} {cell.row}")
                                        rNumSrc = cell.row

                                        dst = load_workbook(filename=str(baseFolder) + '/Sources/' + str(regName) + '/' + str(srcFile) + str(fileExt))
                                        #dst = xlrd.open_workbook('Sources/' + str(regName) + '/' + str(srcFile) + str(fileExt))
                                        q1_dst_ws = dst[qtr]
                                        #q1_dst_ws = dst.sheet_by_name(qtr)

                                        if os.path.isfile(fpath) == True:
                                            for comm in comm_arr:
                                                for srv in srv_arr:
                                                    if comm == "native":
                                                        if srv == " B":
                                                            rNumDst = 22
                                                        elif srv == " C":
                                                            rNumDst = 28
                                                        elif srv == "_T":
                                                            rNumDst = 34
                                                    elif comm == "gamefowl":
                                                        if srv == " B":
                                                            rNumDst = 40
                                                        elif srv == " C":
                                                            rNumDst = 46
                                                        elif srv == "_T":
                                                            rNumDst = 52
                                                    elif comm == "broiler":
                                                        if srv == " B":
                                                            rNumDst = 58
                                                        elif srv == " C":
                                                            rNumDst = 64
                                                        elif srv == "_T":
                                                            rNumDst = 70
                                                    elif comm == "layer":
                                                        if srv == " B":
                                                            rNumDst = 76
                                                        elif srv == " C":
                                                            rNumDst = 82
                                                        elif srv == "_T":
                                                            rNumDst = 88

                                                    sheetNameSrc = str(comm) + str(srv)
                                                    qtr_src_ws = src[sheetNameSrc]
                                                    print(f"PSA-LPSD (EDRW): Loading source -> {province} - {comm} - {srv}")

                                                    if srv != "_T":
                                                        #print(f"Copying Source B{rNumSrc} to E{rNumDst}")
                                                        q1_dst_ws['E' + str(rNumDst)].value = qtr_src_ws['B' + str(rNumSrc)].value
                                                        q1_dst_ws['F' + str(rNumDst)].value = qtr_src_ws['C' + str(rNumSrc)].value
                                                        q1_dst_ws['H' + str(rNumDst)].value = qtr_src_ws['E' + str(rNumSrc)].value
                                                        q1_dst_ws['I' + str(rNumDst)].value = qtr_src_ws['F' + str(rNumSrc)].value
                                                        q1_dst_ws['J' + str(rNumDst)].value = qtr_src_ws['G' + str(rNumSrc)].value
                                                        q1_dst_ws['M' + str(rNumDst)].value = qtr_src_ws['J' + str(rNumSrc)].value
                                                        q1_dst_ws['N' + str(rNumDst)].value = qtr_src_ws['K' + str(rNumSrc)].value
                                                        q1_dst_ws['Q' + str(rNumDst)].value = qtr_src_ws['N' + str(rNumSrc)].value
                                                        q1_dst_ws['R' + str(rNumDst)].value = qtr_src_ws['U' + str(rNumSrc)].value
                                                        q1_dst_ws['S' + str(rNumDst)].value = qtr_src_ws['V' + str(rNumSrc)].value
                                                        q1_dst_ws['V' + str(rNumDst)].value = qtr_src_ws['Y' + str(rNumSrc)].value
                                                        q1_dst_ws['X' + str(rNumDst)].value = qtr_src_ws['AA' + str(rNumSrc)].value
                                                    elif srv == "_T":
                                                        if comm == "native" or comm == "broiler" or comm == "layer":
                                                            q1_dst_ws['AA' + str(rNumDst)].value = qtr_src_ws['AD' + str(rNumSrc)].value
                                                        if comm == "broiler" or comm == "layer":
                                                            q1_dst_ws['AC' + str(rNumDst)].value = qtr_src_ws['AF' + str(rNumSrc)].value
                                                    print(f"PSA-LPSD (EDRW): Copied to destination -> {province} - {comm} - {srv} at row {rNumSrc}")

                                            fpath = str(baseFolder) + '/Final/' + str(regName)
                                            if os.path.isdir(fpath) == False:
                                                os.mkdir(fpath)
                                            finalFile = str(baseFolder) + '/Final/' + str(regName) + '/' + str(commcode) + ' ' + str(province) + '_' + str(yr) + '.xlsx'
                                            print(f"PSA-LPSD (EDRW): Saving to file -> {finalFile}")
                                            dst.save(finalFile)
                                        else:
                                            print(f"File {srcFile}{fileExt} Not Found.")
                        else:
                            print(f"File {srcFile}{fileExt} Not Found.")
                else:
                    print("Region not found or has no provinces.")
            # Duck
            if commcode == '09':
                src_active = src["Backyard"] # Get Province Names from this worksheet
                # comm_arr = ["native","gamefowl","broiler","layer"]
                srv_arr = ["Backyard","Commercial","Total"]
                if selected_provinces:
                    for province in selected_provinces:
                        srcFile = str(commcode) + ' ' + str(province) + '_' + str(yr) # template files
                        fileExt = '.xlsx' # orig .xlsm; .xlsx for openpyxl; .xls for xlrd
                        fpath = str(baseFolder) + '/Sources/' + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(srcFile) + str(fileExt)
                        if os.path.isfile(fpath) == True:
                            for row in src_active.iter_rows(min_row=18, min_col=2, max_row=118, max_col=2):
                                for cell in row:
                                    if cell.value == province:
                                        #print(f"{province} {cell.row}")
                                        rNumSrc = cell.row

                                        #dst = load_workbook(filename=str(baseFolder) + '/Sources/' + str(regName) + '/' + str(srcFile) + str(fileExt))
                                        dst = load_workbook(filename=fpath,keep_links=True)
                                        #dst = xlrd.open_workbook(str(baseFolder) + '/' + 'Sources/' + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(srcFile) + str(fileExt))
                                        #dst = xlrd.open_workbook(fpath)
                                        qtr_dst_ws = dst[qtr]
                                        currqtr = qtr[:-1] + str(int(qtr[-1])+1) # current quarter +1
                                        currqtr_dst_ws = dst[currqtr]
                                        #q1_dst_ws = dst.sheet_by_name(qtr)

                                        if os.path.isfile(fpath) == True:
                                            # for comm in comm_arr:
                                            for srv in srv_arr:
                                                if srv == "Backyard": # row number of destinations of copied values from SD
                                                    rNumDst = 22
                                                elif srv == "Commercial":
                                                    rNumDst = 28
                                                elif srv == "Total":
                                                    rNumDst = 32
                                                # sheetNameSrc = str(comm) + str(srv)
                                                qtr_src_ws = src[srv]
                                                print(f"PSA-LPSD (EDRW): Loading source -> {province} - {srv}")

                                                if srv != "Total":
                                                    #print(f"Copying Source B{rNumSrc} to E{rNumDst}")
                                                    qtr_dst_ws['H' + str(rNumDst)].value = qtr_src_ws['F' + str(rNumSrc)].value # Hatched live
                                                    qtr_dst_ws['J' + str(rNumDst)].value = qtr_src_ws['H' + str(rNumSrc)].value # Laying flock
                                                    qtr_dst_ws['K' + str(rNumDst)].value = qtr_src_ws['I' + str(rNumSrc)].value # MALE BREEDER
                                                    qtr_dst_ws['N' + str(rNumDst)].value = qtr_src_ws['L' + str(rNumSrc)].value # AVIAN INF
                                                    qtr_dst_ws['O' + str(rNumDst)].value = qtr_src_ws['M' + str(rNumSrc)].value # OTHER DISEASE
                                                    qtr_dst_ws['R' + str(rNumDst)].value = qtr_src_ws['P' + str(rNumSrc)].value # DRESSED ON FARM
                                                    qtr_dst_ws['S' + str(rNumDst)].value = qtr_src_ws['T' + str(rNumSrc)].value # Total E.I. (v Current Q)
                                                    qtr_dst_ws['T' + str(rNumDst)].value = qtr_src_ws['U' + str(rNumSrc)].value # Laying flock (v Current Q)
                                                    currqtr_dst_ws['E' + str(rNumDst)].value = qtr_src_ws['T' + str(rNumSrc)].value # Total E.I. (v Next Q) TO B.I. of next Q
                                                    currqtr_dst_ws['F' + str(rNumDst)].value = qtr_src_ws['U' + str(rNumSrc)].value # Laying flock (v Next Q) TO B.I. of next Q
                                                    qtr_dst_ws['W' + str(rNumDst)].value = qtr_src_ws['X' + str(rNumSrc)].value # SOLD LIVE FOR OTHER PURPOSE
                                                    qtr_dst_ws['Y' + str(rNumDst)].value = qtr_src_ws['Z' + str(rNumSrc)].value # AVG LOCAL LIVEWEIGHT
                                                elif srv == "Total":
                                                    qtr_dst_ws['AB' + str(rNumDst)].value = qtr_src_ws['AC' + str(rNumSrc)].value # INFLOW FROM OTHER PRV
                                                    qtr_dst_ws['AD' + str(rNumDst)].value = qtr_src_ws['AE' + str(rNumSrc)].value # SHIPPED-OUT TO OTHER PRV
                                                # put links
                                                currqtr_dst_ws['E21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 2'!S17,0)"
                                                currqtr_dst_ws['F21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 2'!T17,0)"
                                                currqtr_dst_ws['H21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 3'!BL17,0)"
                                                currqtr_dst_ws['J21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 3'!BN17,0)"
                                                currqtr_dst_ws['K21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 3'!BO17,0)"
                                                currqtr_dst_ws['N21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 10'!X18,0)"
                                                currqtr_dst_ws['O21'].value = "=ROUND(SUM('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 10'!Z18,'C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 10'!AB18),0)"
                                                currqtr_dst_ws['R21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 5'!R17,0)"
                                                currqtr_dst_ws['S21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 2'!AI17,0)"
                                                currqtr_dst_ws['T21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 2'!AJ17,0)"
                                                currqtr_dst_ws['V21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 5'!AD17,0)"
                                                currqtr_dst_ws['W21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 9'!AG17,0)"
                                                currqtr_dst_ws['Y21'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 6'!N17"
                                                currqtr_dst_ws['AD21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 7'!T17,0)"
                                                currqtr_dst_ws['E27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 2'!P17,0)"
                                                currqtr_dst_ws['F27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 2'!Q17,0)"
                                                currqtr_dst_ws['H27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 3'!BI17,0)"
                                                currqtr_dst_ws['J27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 3'!BK17,0)"
                                                currqtr_dst_ws['K27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 3'!BL17,0)"
                                                currqtr_dst_ws['N27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 10'!U18,0)"
                                                currqtr_dst_ws['O27'].value = "=ROUND(SUM('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 10'!W18,'C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 10'!Y18),0)"
                                                currqtr_dst_ws['R27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 5'!O17,0)"
                                                currqtr_dst_ws['S27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 2'!AF17,0)"
                                                currqtr_dst_ws['T27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 2'!AG17,0)"
                                                currqtr_dst_ws['V27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 5'!AA17,0)"
                                                currqtr_dst_ws['W27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 9'!AD17,0)"
                                                currqtr_dst_ws['Y27'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 6'!K17"
                                                currqtr_dst_ws['AD27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 7'!Q17,0)"
                                                currqtr_dst_ws['AC31'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[11 " + str(province) + "_23.xlsx]Q4'!$F$17"
                                                currqtr_dst_ws['AC32'].value = "=IF('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[11 " + str(province) + "_23.xlsx]Q3'!Q17=0,'C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[11 " + str(province) + "_23.xlsx]Q2'!X17,'C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[11 " + str(province) + "_23.xlsx]Q3'!Q17)"
                                                print(f"PSA-LPSD (EDRW): Copied to destination -> {province} - {srv} at row {rNumSrc}")
                                            fpath_reg = str(baseFolder) + '/Final/' + str(regCode) + ' ' + str(regName)
                                            if os.path.isdir(fpath_reg) == False:
                                                os.mkdir(fpath_reg)
                                            fpath_prv = str(baseFolder) + '/Final/' + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province)
                                            if os.path.isdir(fpath_prv) == False:
                                                os.mkdir(fpath_prv)
                                            finalFile = fpath_prv + '/' + str(commcode) + ' ' + str(province) + '_' + str(yr) + '.xlsx'
                                            # save links
                                            #qtr_dst_ws['E21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 2'!S17,0)"
                                            
                                            print(f"PSA-LPSD (EDRW): Saving to file -> {finalFile}")
                                            dst.save(finalFile)
                                        else:
                                            print(f"File {srcFile}{fileExt} Not Found.")
                        else:
                            print(f"File {srcFile}{fileExt} Not Found.")
                else:
                    print("Region not found or has no provinces.")
            # Duck Egg
            if commcode == '09a':
                src_active = src["Backyard"] # Get Province Names from this worksheet
                # comm_arr = ["native","gamefowl","broiler","layer"]
                srv_arr = ["Backyard","Commercial","Total"]
                if selected_provinces:
                    for province in selected_provinces:
                        srcFile = str(commcode) + ' ' + str(province) + '_' + str(yr) # template files
                        fileExt = '.xlsx' # orig .xlsm
                        fpath = str(baseFolder) + '/Sources/' + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(srcFile) + str(fileExt)
                        if os.path.isfile(fpath) == True:
                            for row in src_active.iter_rows(min_row=18, min_col=2, max_row=118, max_col=2):
                                for cell in row:
                                    if cell.value == province:
                                        #print(f"{province} {cell.row}")
                                        rNumSrc = cell.row

                                        #dst = load_workbook(filename=str(baseFolder) + '/Sources/' + str(regName) + '/' + str(srcFile) + str(fileExt))
                                        dst = load_workbook(filename=fpath)
                                        #dst = xlrd.open_workbook('Sources/' + str(regName) + '/' + str(srcFile) + str(fileExt))
                                        qtr_dst_ws = dst[qtr]
                                        currqtr = qtr[:-1] + str(int(qtr[-1])+1) # current quarter +1
                                        currqtr_dst_ws = dst[currqtr]
                                        #q1_dst_ws = dst.sheet_by_name(qtr)

                                        if os.path.isfile(fpath) == True:
                                            # for comm in comm_arr:
                                            for srv in srv_arr:
                                                if srv == "Backyard": # row number of destinations of copied values from SD
                                                    rNumDst = 22
                                                elif srv == "Commercial":
                                                    rNumDst = 28
                                                elif srv == "Total":
                                                    rNumDst = 32

                                                # sheetNameSrc = str(comm) + str(srv)
                                                qtr_src_ws = src[srv]
                                                print(f"PSA-LPSD (EDRW): Loading source -> {province} - {srv}")

                                                if srv != "Total":
                                                    #print(f"Copying Source B{rNumSrc} to E{rNumDst}")
                                                    qtr_dst_ws['I' + str(rNumDst)].value = qtr_src_ws['G' + str(rNumSrc)].value # PIECES (EGG PROD)
                                                    qtr_dst_ws['J' + str(rNumDst)].value = qtr_src_ws['H' + str(rNumSrc)].value # CONVERSION FACTOR
                                                    qtr_dst_ws['L' + str(rNumDst)].value = qtr_src_ws['J' + str(rNumSrc)].value # ESTIMATED HATCHING EGGS
                                                elif srv == "Total":
                                                    qtr_dst_ws['P' + str(rNumDst)].value = qtr_src_ws['N' + str(rNumSrc)].value # SHIPPED-OUT TO OTHER PRV
                                                currqtr_dst_ws['E19'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$E$19"
                                                currqtr_dst_ws['F19'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$F$19"
                                                currqtr_dst_ws['E20'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$E$20"
                                                currqtr_dst_ws['F20'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$F$20"
                                                currqtr_dst_ws['E21'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!E21"
                                                currqtr_dst_ws['F21'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!F21"
                                                currqtr_dst_ws['E22'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!E22"
                                                currqtr_dst_ws['F22'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!F22"
                                                currqtr_dst_ws['M19'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$S$19"
                                                currqtr_dst_ws['N19'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$T$19"
                                                currqtr_dst_ws['M20'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$S$20"
                                                currqtr_dst_ws['N20'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$T$20"
                                                currqtr_dst_ws['M21'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!S21"
                                                currqtr_dst_ws['N21'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!T21"
                                                currqtr_dst_ws['M22'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!S22"
                                                currqtr_dst_ws['N22'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!T22"
                                                currqtr_dst_ws['I21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 11'!U17,0)"
                                                currqtr_dst_ws['J21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 11'!AB17,0)"
                                                currqtr_dst_ws['L21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 11'!BN17,0)"
                                                currqtr_dst_ws['P21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Duck " + str(province) + ".xlsx]Table 12'!AG17,0)"
                                                currqtr_dst_ws['H19'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$J$19"
                                                currqtr_dst_ws['H20'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$J$20"
                                                currqtr_dst_ws['H21'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!J21"
                                                currqtr_dst_ws['H22'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!J22"
                                                currqtr_dst_ws['E25'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$E$25"
                                                currqtr_dst_ws['F25'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$F$25"
                                                currqtr_dst_ws['E26'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$E$26"
                                                currqtr_dst_ws['F26'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$F$26"
                                                currqtr_dst_ws['E27'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!E27"
                                                currqtr_dst_ws['F27'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!F27"
                                                currqtr_dst_ws['E28'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!E28"
                                                currqtr_dst_ws['F28'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!F28"
                                                currqtr_dst_ws['M25'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$S$25"
                                                currqtr_dst_ws['N25'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$T$25"
                                                currqtr_dst_ws['M26'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$S$26"
                                                currqtr_dst_ws['N26'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$T$26"
                                                currqtr_dst_ws['M27'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!S27"
                                                currqtr_dst_ws['N27'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!T27"
                                                currqtr_dst_ws['M28'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!S28"
                                                currqtr_dst_ws['N28'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!T28"
                                                currqtr_dst_ws['I27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 11'!R17,0)"
                                                currqtr_dst_ws['J27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 11'!Y17,0)"
                                                currqtr_dst_ws['L27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 11'!BK17,0)"
                                                currqtr_dst_ws['P27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Duck " + str(province) + ".xlsx]Table 12'!AD17,0)"
                                                currqtr_dst_ws['H25'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$J$25"
                                                currqtr_dst_ws['H26'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q3'!$J$26"
                                                currqtr_dst_ws['H27'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!J27"
                                                currqtr_dst_ws['H28'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[09 " + str(province) + "_23.xlsx]Q1'!J28"
                                                print(f"PSA-LPSD (EDRW): Copied to destination -> {province} - {srv} at row {rNumSrc}")
                                            fpath_reg = str(baseFolder) + '/Final/' + str(regCode) + ' ' + str(regName)
                                            if os.path.isdir(fpath_reg) == False:
                                                os.mkdir(fpath_reg)
                                            fpath_prv = str(baseFolder) + '/Final/' + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province)
                                            if os.path.isdir(fpath_prv) == False:
                                                os.mkdir(fpath_prv)
                                            finalFile = fpath_prv + '/' + str(commcode) + ' ' + str(province) + '_' + str(yr) + '.xlsx'
                                            print(f"PSA-LPSD (EDRW): Saving to file -> {finalFile}")
                                            dst.save(finalFile)
                                        else:
                                            print(f"File {srcFile}{fileExt} Not Found.")
                        else:
                            print(f"File {srcFile}{fileExt} Not Found.")
                else:
                    print("Region not found or has no provinces.")
            # Chicken Egg
            if commcode == '08a':
                src_active = src["Native_B"] # Get Province Names from this worksheet
                # comm_arr = ["native","gamefowl","broiler","layer"]
                srv_arr = ["Native_B","Native_C","Broiler_C","Layer_B","Layer_C"]
                if selected_provinces:
                    for province in selected_provinces:
                        srcFile = str(commcode) + ' ' + str(province) + '_' + str(yr) # template files
                        fileExt = '.xlsx' # orig .xlsm
                        fpath = str(baseFolder) + '/Sources/' + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(srcFile) + str(fileExt)
                        if os.path.isfile(fpath) == True:
                            for row in src_active.iter_rows(min_row=18, min_col=2, max_row=118, max_col=2):
                                for cell in row:
                                    if cell.value == province:
                                        #print(f"{province} {cell.row}")
                                        rNumSrc = cell.row

                                        #dst = load_workbook(filename=str(baseFolder) + '/Sources/' + str(regName) + '/' + str(srcFile) + str(fileExt))
                                        dst = load_workbook(filename=fpath)
                                        #dst = xlrd.open_workbook('Sources/' + str(regName) + '/' + str(srcFile) + str(fileExt))
                                        qtr_dst_ws = dst[qtr]
                                        currqtr = qtr[:-1] + str(int(qtr[-1])+1) # current quarter +1
                                        currqtr_dst_ws = dst[currqtr]
                                        #q1_dst_ws = dst.sheet_by_name(qtr)

                                        if os.path.isfile(fpath) == True:
                                            # for comm in comm_arr:
                                            for srv in srv_arr:
                                                if srv == "Native_B": # row number of destinations of copied values from SD
                                                    rNumDst = 22
                                                elif srv == "Native_C":
                                                    rNumDst = 28
                                                elif srv == "Broiler_C":
                                                    rNumDst = 38
                                                elif srv == "Layer_B":
                                                    rNumDst = 44
                                                elif srv == "Layer_C":
                                                    rNumDst = 50

                                                # sheetNameSrc = str(comm) + str(srv)
                                                qtr_src_ws = src[srv]
                                                print(f"PSA-LPSD (EDRW): Loading source -> {province} - {srv}")

                                                if srv == "Native_B" or srv == "Native_C":
                                                    #print(f"Copying Source B{rNumSrc} to E{rNumDst}")
                                                    qtr_dst_ws['J' + str(rNumDst)].value = qtr_src_ws['H' + str(rNumSrc)].value # PIECES (EGG PROD)
                                                    qtr_dst_ws['K' + str(rNumDst)].value = qtr_src_ws['I' + str(rNumSrc)].value # CONVERSION FACTOR
                                                    qtr_dst_ws['M' + str(rNumDst)].value = qtr_src_ws['K' + str(rNumSrc)].value # ESTIMATED HATCHING EGGS
                                                    qtr_dst_ws['Q' + str(rNumDst)].value = qtr_src_ws['O' + str(rNumSrc)].value # SHIPPED-OUT TO OTHER PRV
                                                elif srv == "Broiler_C" or srv == "Layer_C":
                                                    qtr_dst_ws['I' + str(rNumDst)].value = qtr_src_ws['G' + str(rNumSrc)].value # ELER
                                                    qtr_dst_ws['J' + str(rNumDst)].value = qtr_src_ws['H' + str(rNumSrc)].value # PIECES (EGG PROD)
                                                    qtr_dst_ws['K' + str(rNumDst)].value = qtr_src_ws['I' + str(rNumSrc)].value # CONVERSION FACTOR
                                                    qtr_dst_ws['M' + str(rNumDst)].value = qtr_src_ws['K' + str(rNumSrc)].value # ESTIMATED HATCHING EGGS
                                                    qtr_dst_ws['Q' + str(rNumDst)].value = qtr_src_ws['O' + str(rNumSrc)].value # SHIPPED-OUT TO OTHER PRV
                                                elif srv == "Layer_B":
                                                    qtr_dst_ws['I' + str(rNumDst)].value = qtr_src_ws['G' + str(rNumSrc)].value # ELER
                                                    qtr_dst_ws['J' + str(rNumDst)].value = qtr_src_ws['H' + str(rNumSrc)].value # PIECES (EGG PROD)
                                                    qtr_dst_ws['K' + str(rNumDst)].value = qtr_src_ws['I' + str(rNumSrc)].value # CONVERSION FACTOR
                                                    qtr_dst_ws['Q' + str(rNumDst)].value = qtr_src_ws['O' + str(rNumSrc)].value # SHIPPED-OUT TO OTHER PRV
                                                # place links
                                                currqtr_dst_ws['E19'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E19"
                                                currqtr_dst_ws['F19'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F19"
                                                currqtr_dst_ws['E20'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E20"
                                                currqtr_dst_ws['F20'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F20"
                                                currqtr_dst_ws['E21'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E21"
                                                currqtr_dst_ws['F21'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F21"
                                                currqtr_dst_ws['E22'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E22"
                                                currqtr_dst_ws['F22'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F22"
                                                currqtr_dst_ws['N19'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R19"
                                                currqtr_dst_ws['O19'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S19"
                                                currqtr_dst_ws['N20'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R20"
                                                currqtr_dst_ws['O20'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S20"
                                                currqtr_dst_ws['N21'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R21"
                                                currqtr_dst_ws['O21'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S21"
                                                currqtr_dst_ws['N22'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R22"
                                                currqtr_dst_ws['O22'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S22"
                                                currqtr_dst_ws['J21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Native " + str(province) + ".xlsx]Table 9'!L17,0)"
                                                currqtr_dst_ws['K21'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Native " + str(province) + ".xlsx]Table 9'!P17"
                                                currqtr_dst_ws['M21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Native " + str(province) + ".xlsx]Table 9'!BB17,0)"
                                                currqtr_dst_ws['Q21'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Native " + str(province) + ".xlsx]Table 10'!AG17,0)"
                                                currqtr_dst_ws['H19'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I19"
                                                currqtr_dst_ws['H20'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I20"
                                                currqtr_dst_ws['H21'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I21"
                                                currqtr_dst_ws['H22'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I22"
                                                currqtr_dst_ws['E27'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E27"
                                                currqtr_dst_ws['F27'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F27"
                                                currqtr_dst_ws['E28'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E28"
                                                currqtr_dst_ws['F28'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F28"
                                                currqtr_dst_ws['J27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Native " + str(province) + ".xlsx]Table 9'!I17,0)"
                                                currqtr_dst_ws['K27'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Native " + str(province) + ".xlsx]Table 9'!M17"
                                                currqtr_dst_ws['M27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Native " + str(province) + ".xlsx]Table 9'!AY17,0)"
                                                currqtr_dst_ws['N27'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R27"
                                                currqtr_dst_ws['O27'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S27"
                                                currqtr_dst_ws['N28'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R28"
                                                currqtr_dst_ws['O28'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S28"
                                                currqtr_dst_ws['Q27'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Native " + str(province) + ".xlsx]Table 10'!AD17,0)"
                                                currqtr_dst_ws['H27'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I27"
                                                currqtr_dst_ws['H28'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I28"
                                                currqtr_dst_ws['E35'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E61"
                                                currqtr_dst_ws['F35'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F61"
                                                currqtr_dst_ws['E36'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E62"
                                                currqtr_dst_ws['F36'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F62"
                                                currqtr_dst_ws['E37'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E63"
                                                currqtr_dst_ws['F37'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F63"
                                                currqtr_dst_ws['E38'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E64"
                                                currqtr_dst_ws['F38'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F64"
                                                currqtr_dst_ws['N35'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R61"
                                                currqtr_dst_ws['O35'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S61"
                                                currqtr_dst_ws['N36'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R62"
                                                currqtr_dst_ws['O36'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S62"
                                                currqtr_dst_ws['N37'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R63"
                                                currqtr_dst_ws['O37'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S63"
                                                currqtr_dst_ws['N38'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R64"
                                                currqtr_dst_ws['O38'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S64"
                                                currqtr_dst_ws['H36'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I62"
                                                currqtr_dst_ws['I37'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Broiler " + str(province) + ".xlsx]Table 10'!O17"
                                                currqtr_dst_ws['J37'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Broiler " + str(province) + ".xlsx]Table 10'!S17,0)"
                                                currqtr_dst_ws['M37'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Broiler " + str(province) + ".xlsx]Table 10'!BE17,0)"
                                                currqtr_dst_ws['Q37'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Broiler " + str(province) + ".xlsx]Table 11'!AD17,0)"
                                                currqtr_dst_ws['H38'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I64"
                                                currqtr_dst_ws['E41'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E73"
                                                currqtr_dst_ws['F41'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F73"
                                                currqtr_dst_ws['E42'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E74"
                                                currqtr_dst_ws['F42'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F74"
                                                currqtr_dst_ws['E43'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E75"
                                                currqtr_dst_ws['F43'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F75"
                                                currqtr_dst_ws['E44'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E76"
                                                currqtr_dst_ws['F44'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F76"
                                                currqtr_dst_ws['N41'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R73"
                                                currqtr_dst_ws['O41'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S73"
                                                currqtr_dst_ws['N42'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R74"
                                                currqtr_dst_ws['O42'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S74"
                                                currqtr_dst_ws['N43'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R75"
                                                currqtr_dst_ws['O43'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S75"
                                                currqtr_dst_ws['N44'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R76"
                                                currqtr_dst_ws['O44'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S76"
                                                currqtr_dst_ws['I43'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Layer " + str(province) + ".xlsx]Table 8'!R17"
                                                currqtr_dst_ws['J43'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Layer " + str(province) + ".xlsx]Table 8'!V17,0)"
                                                currqtr_dst_ws['K43'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Layer " + str(province) + ".xlsx]Table 8'!Z17"
                                                currqtr_dst_ws['Q43'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWB Layer " + str(province) + ".xlsx]Table 10'!AG17,0)"
                                                currqtr_dst_ws['H41'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I73"
                                                currqtr_dst_ws['H42'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I74"
                                                currqtr_dst_ws['H43'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I75"
                                                currqtr_dst_ws['H44'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I76"
                                                currqtr_dst_ws['E47'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E79"
                                                currqtr_dst_ws['F47'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F79"
                                                currqtr_dst_ws['E48'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E80"
                                                currqtr_dst_ws['F48'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F80"
                                                currqtr_dst_ws['E49'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E81"
                                                currqtr_dst_ws['F49'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F81"
                                                currqtr_dst_ws['E50'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!E82"
                                                currqtr_dst_ws['F50'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!F82"
                                                currqtr_dst_ws['N47'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R79"
                                                currqtr_dst_ws['O47'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S79"
                                                currqtr_dst_ws['N48'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R80"
                                                currqtr_dst_ws['O48'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S80"
                                                currqtr_dst_ws['N49'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R81"
                                                currqtr_dst_ws['O49'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S81"
                                                currqtr_dst_ws['N50'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!R82"
                                                currqtr_dst_ws['O50'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!S82"
                                                currqtr_dst_ws['I49'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Layer " + str(province) + ".xlsx]Table 10'!O17"
                                                currqtr_dst_ws['J49'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Layer " + str(province) + ".xlsx]Table 10'!S17,0)"
                                                currqtr_dst_ws['K49'].value = "='C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Layer " + str(province) + ".xlsx]Table 10'!W17"
                                                currqtr_dst_ws['M49'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Layer " + str(province) + ".xlsx]Table 11'!AQ17,0)"
                                                currqtr_dst_ws['Q49'].value = "=ROUND('C:/EDRW/Q2/Sources/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + str(province) + "_Output Tables/Q3/[Q3 2023_PSWC Layer " + str(province) + ".xlsx]Table 12'!AD17,0)"
                                                currqtr_dst_ws['H47'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I79"
                                                currqtr_dst_ws['H48'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I80"
                                                currqtr_dst_ws['H49'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I81"
                                                currqtr_dst_ws['H50'].value = "='C:/Users/acer/Desktop/EDRW/" + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province) + '/' + "[08 " + str(province) + "_23.xls]Q3'!I82"
                                                print(f"PSA-LPSD (EDRW): Copied to destination -> {province} - {srv} at row {rNumSrc}")
                                            fpath_reg = str(baseFolder) + '/Final/' + str(regCode) + ' ' + str(regName)
                                            if os.path.isdir(fpath_reg) == False:
                                                os.mkdir(fpath_reg)
                                            fpath_prv = str(baseFolder) + '/Final/' + str(regCode) + ' ' + str(regName) + '/' + str(regCode) + ' ' + str(province)
                                            if os.path.isdir(fpath_prv) == False:
                                                os.mkdir(fpath_prv)
                                            finalFile = fpath_prv + '/' + str(commcode) + ' ' + str(province) + '_' + str(yr) + '.xlsx'
                                            print(f"PSA-LPSD (EDRW): Saving to file -> {finalFile}")
                                            dst.save(finalFile)
                                        else:
                                            print(f"File {srcFile}{fileExt} Not Found.")
                        else:
                            print(f"File {srcFile}{fileExt} Not Found.")
                else:
                    print("Region not found or has no provinces.")
            # else:
            #     print(f"Commodity code {commcode} not found.")